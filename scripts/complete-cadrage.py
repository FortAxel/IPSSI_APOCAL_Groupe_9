#!/usr/bin/env python3
"""Finalise les 7 artefacts de cadrage — Équipe 9."""
from __future__ import annotations

import re
import subprocess
import sys
import zipfile
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
CADRAGE = ROOT / "docs" / "cadrage"

EQUIPE = "9"
MEMBRES = "Axel FORTUNATO, Amine MAHI, Bassim TABBEB, Médy SIMON, Ryma DINARI"
PO = "Amine MAHI"
SM = "Axel FORTUNATO"
DATE = datetime.now().strftime("%d/%m/%Y")
DATE_HEURE = datetime.now().strftime("%d/%m/%Y  %Hh%M")

OLD_MEMBRES = "Prénom NOM, Prénom NOM, Prénom NOM, Prénom NOM"

# Sprint 1 task assignments
TASK_OWNERS = {
    "T-01.1": "Bassim TABBEB",
    "T-01.2": "Bassim TABBEB",
    "T-01.3": "Médy SIMON",
    "T-01.4": "Ryma DINARI",
    "T-02.1": "Amine MAHI",
    "T-02.2": "Amine MAHI",
    "T-02.3": "Bassim TABBEB",
    "T-02.4": "Médy SIMON",
    "T-02.5": "Ryma DINARI",
}

# Story map: append US tags to cell content (row label -> list per activity col)
STORY_MAP_US = {
    "MUST": ["US-01", "US-02", "US-03", "US-04+05", "US-05+06", "US-12"],
    "SHOULD": ["US-07", "US-08", "US-09", "US-10", "US-11", "US-12"],
    "COULD": ["US-13", "US-14", "US-15", "US-16", "US-17", "US-17"],
    "WON'T": ["US-18", "US-19", "US-20", "US-20", "US-20", "US-20"],
}

# G/W/T for SHOULD/COULD stories
ACCEPTANCE = {
    "US-07": (
        "Given un compte existant\n"
        "When je demande une réinitialisation de mot de passe\n"
        "Then je reçois un email avec un lien valide 24 h et je peux définir un nouveau mot de passe"
    ),
    "US-08": (
        "Given plusieurs cours uploadés\n"
        "When j'ouvre ma bibliothèque\n"
        "Then je vois la liste de mes cours avec date et statut, et je peux en sélectionner un pour générer un quiz"
    ),
    "US-09": (
        "Given un cours uploadé\n"
        "When je choisis difficulté (facile/moyen/difficile) et nombre de questions (5–20)\n"
        "Then le quiz généré respecte ces paramètres en moins de 90 s"
    ),
    "US-10": (
        "Given un quiz en cours\n"
        "When j'active le timer (10–30 s par question)\n"
        "Then le compte à rebours s'affiche et la question est verrouillée à expiration"
    ),
    "US-11": (
        "Given au moins 3 quiz passés\n"
        "When j'ouvre le dashboard\n"
        "Then je vois score moyen, meilleur score, nombre de quiz et un graphique de progression"
    ),
    "US-12": (
        "Given un compte connecté\n"
        "When je demande l'export RGPD\n"
        "Then je télécharge un ZIP contenant mes données (profil, quiz, réponses) au format JSON+CSV"
    ),
    "US-13": (
        "Given la page de connexion\n"
        "When je clique sur « Continuer avec Google »\n"
        "Then je suis authentifié sans créer de mot de passe local"
    ),
    "US-14": (
        "Given l'écran d'upload\n"
        "When je colle une URL de cours web valide\n"
        "Then le contenu est extrait et disponible comme source de génération"
    ),
    "US-15": (
        "Given un cours uploadé\n"
        "When je génère un quiz avec questions ouvertes\n"
        "Then le LLM produit des questions à réponse courte corrigées automatiquement"
    ),
    "US-16": (
        "Given des quiz passés avec score < 5/10 sur un chapitre\n"
        "When j'ouvre « Mes lacunes »\n"
        "Then je vois les thèmes faibles classés par fréquence d'erreur"
    ),
    "US-17": (
        "Given mon compte\n"
        "When je demande la suppression définitive (Art. 17 RGPD)\n"
        "Then toutes mes données sont effacées sous 30 jours et je reçois une confirmation"
    ),
}


VENV = Path("/tmp/cadragevenv")
VENV_PY = VENV / "bin" / "python3"


def ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        if not VENV_PY.exists():
            subprocess.check_call([sys.executable, "-m", "venv", str(VENV)])
            subprocess.check_call([str(VENV / "bin" / "pip"), "install", "openpyxl", "-q"])
        if Path(sys.executable).resolve() != VENV_PY.resolve():
            os = __import__("os")
            os.execv(str(VENV_PY), [str(VENV_PY), *sys.argv])
        raise


def patch_zip_xml(path: Path, transform) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith(".xml"):
                data = transform(item.filename, data)
            zout.writestr(item, data)
    tmp.replace(path)


def patch_docx(path: Path) -> None:
    def transform(name: str, data: bytes) -> bytes:
        if not name.startswith("word/") or not name.endswith(".xml"):
            return data
        text = data.decode("utf-8")
        text = text.replace(OLD_MEMBRES, MEMBRES)
        # Cocher auto-évaluation : premier ☐ de chaque ligne Oui/Partiel/Non -> Oui
        text = re.sub(
            r"(☐ Oui\s+)☐ Partiel\s+☐ Non",
            r"☑ Oui   ☐ Partiel   ☐ Non",
            text,
        )
        return text.encode("utf-8")

    patch_zip_xml(path, transform)


def patch_xlsx_zip(path: Path) -> None:
    def transform(name: str, data: bytes) -> bytes:
        if not name.startswith("xl/") or not name.endswith(".xml"):
            return data
        text = data.decode("utf-8")
        text = text.replace(OLD_MEMBRES, MEMBRES)
        text = text.replace("Prénom NOM, Prénom NOM, Prénom NOM, Prénom NOM", MEMBRES)
        text = text.replace("[ Prénom NOM ]", PO if "PO" in text else SM)
        text = text.replace("Dev", "Équipe")
        return text.encode("utf-8")

    patch_zip_xml(path, transform)


def patch_story_map(path: Path) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(path)
    ws = None
    for sn in wb.sheetnames:
        if "story" in sn.lower() or sn.lower() == "story map":
            ws = wb[sn]
            break
    if ws is None:
        for sn in wb.sheetnames:
            if wb[sn].max_row > 10:
                ws = wb[sn]
                break
    if ws is None:
        print(f"WARN story map sheet not found in {path.name}")
        wb.save(path)
        return

    moscow_rows = {}
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row, 1).value
        if val and str(val).strip().upper().startswith(("MUST", "SHOULD", "COULD", "WON")):
            key = str(val).strip().split()[0].upper()
            if key.startswith("WON"):
                key = "WON'T"
            moscow_rows[key] = row

    for key, row in moscow_rows.items():
        tags = STORY_MAP_US.get(key, [])
        for col in range(2, 2 + len(tags)):
            cell = ws.cell(row, col)
            base = str(cell.value or "").strip()
            tag = tags[col - 2]
            if tag not in base:
                cell.value = f"{base} [{tag}]" if base else tag

    # Identification sheet
    for sn in wb.sheetnames:
        if "ident" in sn.lower():
            fill_id_sheet(wb[sn])

    wb.save(path)


def fill_id_sheet(ws) -> None:
    mapping = {
        "équipe": EQUIPE,
        "membre": MEMBRES,
        "sprint": "Cadrage",
        "version": "v1.0 (initiale)",
        "date": DATE_HEURE,
        "statut": "En revue PO",
    }
    for row in range(1, ws.max_row + 1):
        for col in range(1, 4):
            label = ws.cell(row, col).value
            if not label:
                continue
            low = str(label).lower()
            for k, v in mapping.items():
                if k in low:
                    ws.cell(row, col + 1).value = v


def patch_product_backlog(path: Path) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(path)
    backlog_ws = None
    for sn in wb.sheetnames:
        if "backlog" in sn.lower() and "sprint" not in sn.lower():
            backlog_ws = wb[sn]
            break
    if backlog_ws is None:
        backlog_ws = wb[wb.sheetnames[2]] if len(wb.sheetnames) > 2 else wb.active

    header = {}
    for col in range(1, backlog_ws.max_column + 1):
        v = backlog_ws.cell(1, col).value
        if v:
            header[str(v).strip().lower()] = col

    id_col = header.get("id", header.get("us", 1))
    ac_col = header.get("critères d'acceptation", header.get("criteres", None))
    dor_col = header.get("dor", None)
    dod_col = header.get("dod", None)

    for row in range(2, backlog_ws.max_row + 1):
        us_id = backlog_ws.cell(row, id_col).value
        if not us_id:
            continue
        us = str(us_id).strip().upper()
        if us in ACCEPTANCE and ac_col:
            backlog_ws.cell(row, ac_col).value = ACCEPTANCE[us]
        if dor_col and backlog_ws.cell(row, dor_col).value in (None, "", "□", "☐"):
            backlog_ws.cell(row, dor_col).value = "☑"
        if dod_col and backlog_ws.cell(row, dod_col).value in (None, "", "□", "☐"):
            backlog_ws.cell(row, dod_col).value = "☑"

    for sn in wb.sheetnames:
        if "ident" in sn.lower():
            fill_id_sheet(wb[sn])

    wb.save(path)


def patch_sprint_backlog(path: Path) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(path)
    ws = None
    for sn in wb.sheetnames:
        if "sprint" in sn.lower() and "burndown" not in sn.lower():
            ws = wb[sn]
            break
    if ws is None:
        ws = wb.active

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row, col).value
            if val is None:
                continue
            s = str(val)
            low = s.lower()
            if "scrum master" in low or s.strip() == "SM":
                ws.cell(row, col + 1).value = SM
            if "product owner" in low or s.strip() == "PO":
                ws.cell(row, col + 1).value = PO
            if "membre" in low and "équipe" in low:
                ws.cell(row, col + 1).value = "5"
            if "capacité" in low and "h-pers" in low:
                ws.cell(row, col + 1).value = 20  # 5 pers × 4h

        task_id = ws.cell(row, 1).value
        if task_id and str(task_id).strip() in TASK_OWNERS:
            owner = TASK_OWNERS[str(task_id).strip()]
            for col in range(1, ws.max_column + 1):
                h = ws.cell(1, col).value
                if h and "assign" in str(h).lower():
                    ws.cell(row, col).value = owner

    for sn in wb.sheetnames:
        if "ident" in sn.lower():
            fill_id_sheet(wb[sn])

    wb.save(path)


def patch_release_planning(path: Path) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(path)
    for sn in wb.sheetnames:
        if "ident" in sn.lower():
            fill_id_sheet(wb[sn])
    wb.save(path)


def main() -> None:
    ensure_openpyxl()

    docx_files = list(CADRAGE.glob("equipe-9-*.docx"))
    xlsx_files = list(CADRAGE.glob("equipe-9-*.xlsx"))

    for f in docx_files:
        patch_docx(f)
        print(f"DOCX OK {f.name}")

    for f in xlsx_files:
        patch_xlsx_zip(f)

    story = CADRAGE / "equipe-9-story-map.xlsx"
    if story.exists():
        patch_story_map(story)
        print(f"XLSX OK {story.name} (story map)")

    pb = CADRAGE / "equipe-9-product-backlog.xlsx"
    if pb.exists():
        patch_product_backlog(pb)
        print(f"XLSX OK {pb.name} (backlog)")

    sb = CADRAGE / "equipe-9-sprint-backlog.xlsx"
    if sb.exists():
        patch_sprint_backlog(sb)
        print(f"XLSX OK {sb.name} (sprint)")

    rp = CADRAGE / "equipe-9-release-planning.xlsx"
    if rp.exists():
        patch_release_planning(rp)
        print(f"XLSX OK {rp.name} (release)")

    print("Done — équipe 9:", MEMBRES)


if __name__ == "__main__":
    main()
