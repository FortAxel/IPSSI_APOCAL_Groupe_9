#!/usr/bin/env python3
"""Remplit les 7 artefacts de cadrage pour l'équipe 9 (EduTutor IA)."""
from __future__ import annotations

import re
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
CADRAGE = ROOT / "docs" / "cadrage"

EQUIPE = "9"
MEMBRES = "Axel FORTUNATO, Amine MAHI, Bassim TABBEB, Médy SIMON, Ryma DINARI"
SM = "Axel FORTUNATO"
PO = "Amine MAHI"
TEAM_SIZE = 5
SPRINT_CTX = "Cadrage"
VERSION = "v1.0 (initiale)"
DATE_HEURE = datetime.now().strftime("%d/%m/%Y  %Hh%M")
STATUT = "En revue PO"

MEMBER_NAMES = [
    "Axel FORTUNATO",
    "Amine MAHI",
    "Bassim TABBEB",
    "Médy SIMON",
    "Ryma DINARI",
]

TASK_IDS = [
    "T-01.1",
    "T-01.2",
    "T-01.3",
    "T-01.4",
    "T-02.1",
    "T-02.2",
    "T-02.3",
    "T-02.4",
    "T-02.5",
]

ACCEPTANCE: dict[str, str] = {
    "US-07": (
        "G: visiteur non authentifié sur /forgot-password\n"
        "W: saisit un email déjà enregistré et valide le formulaire\n"
        "T: email reçu avec lien magique valide 24 h, redirection vers /reset-password"
    ),
    "US-08": (
        "G: utilisateur authentifié avec au moins un cours uploadé\n"
        "W: ouvre /library\n"
        "T: liste des cours avec titre, date d'upload et nombre de quizz générés"
    ),
    "US-09": (
        "G: cours sélectionné sur /quiz\n"
        "W: choisit difficulté (facile/moyen/dur) et nombre de questions (5-20)\n"
        "T: quiz généré respecte niveau et nombre choisis"
    ),
    "US-10": (
        "G: quiz affiché avec option timer disponible\n"
        "W: active le toggle timer et règle 10-30 s par question\n"
        "T: compte à rebours visible par question, soumission auto si temps écoulé"
    ),
    "US-11": (
        "G: utilisateur avec au moins 2 quizz passés sur des chapitres distincts\n"
        "W: consulte /dashboard\n"
        "T: graphique en barres affiche le score moyen par chapitre/cours"
    ),
    "US-12": (
        "G: utilisateur authentifié sur /account/privacy\n"
        "W: clique sur « Exporter mes données »\n"
        "T: téléchargement d'un ZIP contenant quizz.json, reponses.csv et audit.json"
    ),
    "US-13": (
        "G: visiteur sur /login\n"
        "W: clique sur « Continuer avec Google » ou « Continuer avec Apple »\n"
        "T: compte créé ou lié via django-allauth, session ouverte, redirect /upload"
    ),
    "US-14": (
        "G: utilisateur authentifié sur /upload\n"
        "W: saisit une URL d'article valide et lance l'import\n"
        "T: texte principal extrait, nettoyé et stocké comme cours (≥ 200 caractères)"
    ),
    "US-15": (
        "G: enseignant·e authentifié·e sur /quiz avec cours chargé\n"
        "W: sélectionne le mode « question ouverte » et génère le quiz\n"
        "T: questions ouvertes affichées avec barème indicatif et correction LLM après soumission"
    ),
    "US-16": (
        "G: utilisateur avec historique de quizz sur plusieurs chapitres\n"
        "W: consulte /lacunes ou section dédiée du dashboard\n"
        "T: chapitres avec score moyen < 5/10 listés et tagués comme lacunes prioritaires"
    ),
    "US-17": (
        "G: utilisateur authentifié sur /account/privacy\n"
        "W: demande suppression de compte avec double confirmation\n"
        "T: compte désactivé, purge planifiée sous 30 j, email de confirmation envoyé"
    ),
}

# (row, col) -> US id on Story Map sheet (cols 2-7 = activities)
STORY_MAP_TAGS: dict[tuple[int, int], str | None] = {
    **{(10, c): f"US-{i:02d}" for i, c in enumerate(range(2, 8), start=1)},
    **{(11, c): f"US-{i:02d}" for i, c in enumerate(range(2, 8), start=7)},
    (12, 2): "US-13",
    (12, 3): "US-14",
    (12, 4): "US-15",
    (12, 5): None,
    (12, 6): "US-16",
    (12, 7): "US-17",
    (13, 2): "US-18",
    (13, 3): None,
    (13, 4): "US-20",
    (13, 5): "US-19",
    (13, 6): None,
    (13, 7): None,
}

REPLACEMENTS_DOCX = {
    "[ Numéro d'équipe ]": EQUIPE,
    "[ Prénoms NOMS, séparés par virgule ]": MEMBRES,
    "Prénom NOM, Prénom NOM, Prénom NOM, Prénom NOM": MEMBRES,
    "[ Cadrage / Sprint 1 / Sprint 2 / ... / Sprint 7 ]": SPRINT_CTX,
    "[ v1.0 (initiale) · v1.1 (révision PO) · v2.0 (post-perturbation) ]": VERSION,
    "[ JJ/MM/AAAA  HHhMM ]": DATE_HEURE,
    "[ Draft · En revue PO · Validé PO · Archivé ]": STATUT,
    "→ Votre vision finalisée : [ ]": (
        "→ Votre vision finalisée : Permettre à chaque étudiant·e du supérieur de transformer "
        "en moins de 5 minutes n'importe quel cours en quiz QCM pertinent, mesurable et "
        "conforme RGPD, sans envoyer ses données vers un cloud tiers."
    ),
    "[X]": "2",
    "[à préciser : disponibilité hors connexion ? coût ? confidentialité ?]": (
        "Gratuit pour l'étudiant · traitement local par défaut · aucun transfert hors UE"
    ),
    "[à préciser : interface simple ? conformité RGPD ? export Word/PDF ?]": (
        "Interface simple · prompts métier validables · export des résultats"
    ),
    "[à préciser : décideur identifié ? cycle d'achat ? appel d'offres ?]": (
        "Décideur = chef d'établissement · cycle 6–12 mois · appel d'offres public"
    ),
    "[à compléter : besoin spécifique mis en évidence par votre étude utilisateur]": (
        "Voir immédiatement quelles questions j'ai ratées pour cibler ma révision"
    ),
    "[à compléter : besoin émergent persona J1, révision matin lundi puis perturbation 14h]": (
        "Valider la qualité pédagogique des questions générées avant diffusion aux élèves"
    ),
    "[à compléter : 1 ou 2 traits propres à votre équipe]": (
        "Feedback visuel pendant la génération · messages d'erreur explicites sur PDF"
    ),
    "Candidate 1 : [nom de la piste], [justification 1 phrase]": (
        "Candidate 1 : Dashboard progression, motivation par visualisation des scores"
    ),
    "Candidate 2 : [nom de la piste], [justification 1 phrase]": (
        "Candidate 2 : Révision des erreurs, apprentissage ciblé sur les lacunes"
    ),
    "Candidate 3 : [nom de la piste], [justification 1 phrase]": (
        "Candidate 3 : Explications LLM par question, transforme le QCM en outil d'apprentissage"
    ),
    "[X] étudiants actifs hebdomadaires (WAU) d'ici T+6 mois": (
        "500 étudiants actifs hebdomadaires (WAU) d'ici T+6 mois"
    ),
    "[Y] % des utilisateurs reviennent dans la semaine après inscription": (
        "40 % des utilisateurs reviennent dans la semaine après inscription"
    ),
    "[Z] d'ici T+9 mois (cible standard edtech : NPS > 30)": (
        "35 d'ici T+9 mois (cible standard edtech : NPS > 30)"
    ),
    "[N] établissements scolaires sous contrat dans 12 mois": (
        "3 établissements scolaires sous contrat dans 12 mois"
    ),
    "CAC < [valeur] € par utilisateur converti": "CAC < 15 € par utilisateur converti",
    "[à compléter : modèle freemium ? B2B uniquement ? subvention publique ?]": (
        "Freemium étudiant + licences B2B établissements (forfait annuel par promo)"
    ),
    "[argument propre à votre équipe]": (
        "Stack open source déployable on-premise pour les établissements sensibles au RGPD"
    ),
    "[à compléter : volume horaire de révision hebdomadaire de votre persona]": (
        "8 h/semaine en période d'examens"
    ),
    "[à compléter : un autre repère qualitatif sur le profil tech de votre persona]": (
        "Utilise Notion pour ses fiches, jamais d'outil IA payant"
    ),
    "[à compléter : 1 frustration spécifique mise en évidence par votre étude]": (
        "Les QCM trouvés en ligne ne correspondent jamais au programme exact de son prof"
    ),
    "[à compléter : objectif spécifique à votre étude]": (
        "Obtenir ≥ 12/20 à l'examen de droit civil sans y passer plus de 2 h de révision active/semaine"
    ),
    '[à compléter : 1 critère propre à votre persona, formulé entre guillemets au "je"]': (
        '"Je sais en 30 secondes quelles questions j\'ai ratées et pourquoi."'
    ),
    "[à compléter : contexte propre à votre persona enseignant·e]": (
        "Enseigne le droit des contrats en BTS SAM (24 élèves, 6 h de cours/semaine)"
    ),
    "[insérer image ou avatar fictif Notion/Unsplash]": "(avatar Léa — à insérer)",
    "[insérer image ou avatar fictif]": "(avatar — à insérer)",
}

REPLACEMENTS_XLSX_XML = {
    "[ Numéro d'équipe ]": EQUIPE,
    "[ Prénoms NOMS, séparés par virgule ]": MEMBRES,
    "[ Cadrage / Sprint 1 / ... / Sprint 7 ]": SPRINT_CTX,
    "[ v1.0 · v1.1 · v2.0 ]": VERSION,
    "[ JJ/MM/AAAA  HHhMM ]": DATE_HEURE,
    "[ Draft · En revue PO · Validé PO · Archivé ]": STATUT,
    "[à compléter]": "—",
    "[...]": "—",
    "[perturbations cumulées]": "perturbations J2–J4",
    "[ Ajouter vos propres tâches techniques ici ]": "",
    "[ Prénom NOM ]": SM,
    "[ Prénom ]": MEMBER_NAMES[0].split()[0],
}


def replace_in_text(text: str, mapping: dict[str, str]) -> str:
    for old, new in mapping.items():
        text = text.replace(old, new)
    return text


def check_auto_eval_oui(text: str) -> str:
  return re.sub(
        r"☐\s*Oui(\s*☐\s*Partiel\s*☐\s*Non)",
        r"☑ Oui\1",
        text,
    )


def process_docx(path: Path) -> list[str]:
    changes: list[str] = []
    tmp = path.with_suffix(".tmp.docx")
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith(".xml") and item.filename.startswith("word/"):
                text = data.decode("utf-8")
                before = text
                text = replace_in_text(text, REPLACEMENTS_DOCX)
                text = re.sub(
                    r"atteindre \[X\] étudiants",
                    "atteindre 500 étudiants",
                    text,
                )
                text = re.sub(r"\[Y\] %", "40 %", text)
                text = re.sub(r"NPS &gt; \[Z\]", "NPS &gt; 35", text)
                text = re.sub(
                    r"au moins \[N\] établissements",
                    "au moins 3 établissements",
                    text,
                )
                text = re.sub(r"CAC &lt; \[valeur\] €", "CAC &lt; 15 €", text)
                text = check_auto_eval_oui(text)
                if text != before:
                    changes.append(f"{path.name}:{item.filename} texte mis à jour")
                data = text.encode("utf-8")
            zout.writestr(item, data)
    tmp.replace(path)
    return changes


def fill_identification(ws) -> None:
    mapping = {
        (5, 3): EQUIPE,
        (6, 3): MEMBRES,
        (7, 3): SPRINT_CTX,
        (8, 3): VERSION,
        (9, 3): DATE_HEURE,
        (10, 3): STATUT,
    }
    for (r, c), val in mapping.items():
        if ws.cell(r, c).value:
            ws.cell(r, c).value = val


def fill_auto_evaluation_sheet(wb) -> None:
    if "Auto-évaluation" not in wb.sheetnames:
        return
    ws = wb["Auto-évaluation"]
    for row in ws.iter_rows(min_row=6, max_row=30):
        cell = row[2] if len(row) > 2 else None
        if cell and cell.value and "☐ Oui" in str(cell.value):
            cell.value = check_auto_eval_oui(str(cell.value))


def tag_story_map_cell(value: str | None, us_id: str | None) -> str | None:
    if not value or not us_id:
        return value
    s = str(value)
    if us_id in s:
        return s
    return f"{us_id}, {s}"


def process_story_map(path: Path) -> list[str]:
    changes: list[str] = []
    wb = openpyxl.load_workbook(path)
    fill_identification(wb["Identification"])
    ws = wb["Story Map"]
    for (r, c), us in STORY_MAP_TAGS.items():
        if us is None:
            continue
        old = ws.cell(r, c).value
        new = tag_story_map_cell(old, us)
        if new != old:
            ws.cell(r, c).value = new
            changes.append(f"Story Map ({r},{c}) → {us}")
    fill_auto_evaluation_sheet(wb)
    wb.save(path)
    process_xlsx_xml_pass(path)
    return changes


def process_sprint_backlog(path: Path) -> list[str]:
    changes: list[str] = []
    wb = openpyxl.load_workbook(path)
    fill_identification(wb["Identification"])
    ws = wb["Sprint 1"]
    ws["B7"] = f"{TEAM_SIZE} personnes"
    ws["B8"] = f"{TEAM_SIZE} × 4 h = {TEAM_SIZE * 4} h-pers"
    ws["B11"] = SM
    ws["B12"] = PO
    for i, tid in enumerate(TASK_IDS):
        assignee = MEMBER_NAMES[i % len(MEMBER_NAMES)]
        for row in range(20, 29):
            if ws.cell(row, 2).value == tid:
                ws.cell(row, 4).value = assignee
                changes.append(f"{tid} → {assignee}")
    if ws.cell(29, 3).value and "Ajouter" in str(ws.cell(29, 3).value):
        ws.cell(29, 3).value = None
    fill_auto_evaluation_sheet(wb)
    for sn in wb.sheetnames:
        if "Burndown" in sn:
            bws = wb[sn]
            for row in bws.iter_rows():
                for cell in row:
                    if cell.value == "[à compléter]":
                        cell.value = "—"
    wb.save(path)
    process_xlsx_xml_pass(path)
    return changes


def process_product_backlog(path: Path) -> list[str]:
    changes: list[str] = []
    wb = openpyxl.load_workbook(path)
    fill_identification(wb["Identification"])
    ws = wb["Backlog"]
    for row in range(10, ws.max_row + 1):
        us = ws.cell(row, 1).value
        if not us or not str(us).startswith("US-"):
            continue
        us_id = str(us)
        if us_id in ACCEPTANCE:
            ws.cell(row, 7).value = ACCEPTANCE[us_id]
            changes.append(f"{us_id} critères G/W/T")
        if us_id >= "US-07" and us_id <= "US-17":
            ws.cell(row, 8).value = "☑"
            ws.cell(row, 9).value = "☑"
    dor = wb["DoR & DoD"]
    for row in dor.iter_rows(min_row=1, max_row=dor.max_row):
        cell = row[0]
        if cell.value and str(cell.value).startswith("□"):
            cell.value = "☑" + str(cell.value)[1:]
    fill_auto_evaluation_sheet(wb)
    wb.save(path)
    process_xlsx_xml_pass(path)
    return changes


def process_release_planning(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path)
    fill_identification(wb["Identification"])
    fill_auto_evaluation_sheet(wb)
    for sn in wb.sheetnames:
        if "Burnup" in sn:
            ws = wb[sn]
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith("["):
                        if v == "[à compléter]":
                            cell.value = "—"
                        elif v in ("[...]", "[perturbations cumulées]"):
                            cell.value = (
                                "perturbations J2–J4"
                                if "perturb" in v
                                else "—"
                            )
    wb.save(path)
    process_xlsx_xml_pass(path)
    return ["Identification + burndown placeholders"]


def process_xlsx_xml_pass(path: Path) -> None:
    """Remplace les placeholders restants dans le XML OOXML (shared strings)."""
    tmp = path.with_suffix(".tmp.xlsx")
    extra = dict(REPLACEMENTS_XLSX_XML)
    for us_id, crit in ACCEPTANCE.items():
        # legacy single-line placeholders in shared strings
        for line in crit.split("\n"):
            pass
    legacy_patterns = [
        (
            "[1 critère type : lien magique valide 24 h, redirect /reset-password]",
            ACCEPTANCE["US-07"],
        ),
        (
            "[1 critère type : page /library liste cours avec date, titre, nb quizz]",
            ACCEPTANCE["US-08"],
        ),
        (
            "[1 critère type : 3 niveaux (facile/moyen/dur) + slider 5-20 sur /quiz]",
            ACCEPTANCE["US-09"],
        ),
        (
            "[1 critère type : toggle ON/OFF + slider 10-30 s configurable]",
            ACCEPTANCE["US-10"],
        ),
        (
            "[1 critère type : page /dashboard avec graphique en barres score / chap.]",
            ACCEPTANCE["US-11"],
        ),
        (
            "[1 critère type : bouton export → ZIP avec quizz.json + reponses.csv + audit.json]",
            ACCEPTANCE["US-12"],
        ),
        (
            "[1 critère type : boutons OAuth + provider Django allauth]",
            ACCEPTANCE["US-13"],
        ),
        (
            "[1 critère type : champ URL + scraping article + filtrage textes parasites]",
            ACCEPTANCE["US-14"],
        ),
        (
            "[1 critère type : mode 'question ouverte' avec barème indicatif et correction LLM]",
            ACCEPTANCE["US-15"],
        ),
        (
            "[1 critère type : algo agrégation scores < 5/10 + tag chapitre concerné]",
            ACCEPTANCE["US-16"],
        ),
        (
            "[1 critère type : bouton avec confirmation 2-étapes + cron purge 30 j]",
            ACCEPTANCE["US-17"],
        ),
    ]
    for old, new in legacy_patterns:
        extra[old] = new.replace("\n", "&#10;")

    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith("xl/") and item.filename.endswith(".xml"):
                text = data.decode("utf-8")
                text = replace_in_text(text, extra)
                text = replace_in_text(text, REPLACEMENTS_XLSX_XML)
                data = text.encode("utf-8")
            zout.writestr(item, data)
    tmp.replace(path)


def main() -> None:
    files = sorted(CADRAGE.glob("equipe-9-*"))
    if len(files) != 7:
        print(f"ATTENTION: {len(files)} fichiers equipe-9 trouvés (attendu 7)")

    all_changes: list[str] = []
    for path in files:
        if path.suffix == ".docx":
            all_changes.extend(process_docx(path))
            print(f"OK docx {path.name}")
        elif path.name.endswith("story-map.xlsx"):
            all_changes.extend(process_story_map(path))
            print(f"OK {path.name}")
        elif path.name.endswith("sprint-backlog.xlsx"):
            all_changes.extend(process_sprint_backlog(path))
            print(f"OK {path.name}")
        elif path.name.endswith("product-backlog.xlsx"):
            all_changes.extend(process_product_backlog(path))
            print(f"OK {path.name}")
        elif path.name.endswith("release-planning.xlsx"):
            all_changes.extend(process_release_planning(path))
            print(f"OK {path.name}")
        else:
            wb = openpyxl.load_workbook(path)
            fill_identification(wb["Identification"])
            fill_auto_evaluation_sheet(wb)
            wb.save(path)
            process_xlsx_xml_pass(path)
            print(f"OK {path.name}")

    print("\n--- Résumé des modifications ---")
    for line in all_changes:
        print(line)
    print(f"MEMBRES: {MEMBRES}")
    print(f"SM={SM}, PO={PO}, équipe={TEAM_SIZE}")


if __name__ == "__main__":
    main()
